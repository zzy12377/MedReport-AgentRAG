# -*- coding: utf-8 -*-
"""
prepare_ddxplus_for_medrag.py

把 DDXPlus 原始数据转换成 SNOWTEAM2023/MedRAG 项目更容易读取的格式。

默认输入目录：
    dataset/ddxplus_raw/

默认输出：
    dataset/AI Data Set with Categories.csv
    dataset/df/train/participant_1.json ...
    dataset/df/test/participant_1.json ...

第一次跑通建议直接运行：
    python scripts/prepare_ddxplus_for_medrag.py

如果你想全量生成：
    python scripts/prepare_ddxplus_for_medrag.py --all

如果只想少量测试：
    python scripts/prepare_ddxplus_for_medrag.py --max-train 50 --max-test 20 --max-validate 20
"""

from __future__ import annotations

import argparse
import ast
import csv
import io
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple


OUTPUT_CSV_COLUMNS = [
    "Participant No.",
    "Age",
    "Sex",
    "Pain Presentation and Description",
    "Pain descriptions and assorted symptoms (self-report)",
    "Pain restriction",
    "Processed Diagnosis",
    "Level 2",
    "Level 1",
]


def repo_root() -> Path:
    # 当前脚本应放在 MedRAG-main/scripts/prepare_ddxplus_for_medrag.py
    # parents[1] 就是 MedRAG-main
    return Path(__file__).resolve().parents[1]


def norm_key(s: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s or "").strip().lower())


def clean_text(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s




def text_from_any(value: Any) -> str:
    """
    从 DDXPlus 可能出现的多语言 dict / list / 普通字符串中提取英文文本。
    例如：
        {"fr": "sensible", "en": "sensitive"} -> sensitive
        [{"en": "a cough"}] -> a cough
    """
    if value is None:
        return ""

    if isinstance(value, str):
        s = clean_text(value)
        # 字符串里可能本身是 "{'fr': ..., 'en': ...}"
        if (s.startswith("{") and s.endswith("}")) or (s.startswith("[") and s.endswith("]")):
            try:
                obj = ast.literal_eval(s)
                extracted = text_from_any(obj)
                if extracted:
                    return extracted
            except Exception:
                pass
        return s

    if isinstance(value, (int, float)):
        return clean_text(value)

    if isinstance(value, dict):
        # 优先英文
        for k in ["en", "eng", "english", "text_en", "label_en", "question_en", "name_en"]:
            if k in value and value[k] not in (None, ""):
                return text_from_any(value[k])

        # 再找常见文本字段
        for k in [
            "name",
            "label",
            "display_name",
            "question",
            "question_text",
            "description",
            "meaning",
            "text",
            "value",
        ]:
            if k in value and value[k] not in (None, ""):
                return text_from_any(value[k])

        # 再兜底找任意包含 en 的 key
        for k, v in value.items():
            if "en" in str(k).lower() and v not in (None, ""):
                return text_from_any(v)

        # 最后取第一个非空字符串
        for _, v in value.items():
            extracted = text_from_any(v)
            if extracted:
                return extracted

        return ""

    if isinstance(value, list):
        parts = [text_from_any(x) for x in value]
        parts = [p for p in parts if p]
        return "; ".join(parts)

    return clean_text(value)


def get_field(row: Dict[str, Any], names: List[str], default: str = "") -> str:
    """兼容大小写和下划线差异读取字段。"""
    if not row:
        return default

    direct = {}
    normalized = {}
    for k, v in row.items():
        direct[str(k)] = v
        normalized[norm_key(k)] = v

    for name in names:
        if name in direct:
            return clean_text(direct[name])
        nk = norm_key(name)
        if nk in normalized:
            return clean_text(normalized[nk])

    return default


def parse_maybe_list(value: Any) -> List[Any]:
    """把 DDXPlus 里类似 "['E_1', 'E_2']" 或 JSON 字符串转成 list。"""
    if value is None:
        return []

    if isinstance(value, list):
        return value

    s = clean_text(value)
    if not s:
        return []

    # 先试 JSON
    try:
        obj = json.loads(s)
        if isinstance(obj, list):
            return obj
        return [obj]
    except Exception:
        pass

    # 再试 Python 字面量，DDXPlus CSV 很可能是这种
    try:
        obj = ast.literal_eval(s)
        if isinstance(obj, list):
            return obj
        return [obj]
    except Exception:
        pass

    # 最后兜底：按常见分隔符切
    if ";" in s:
        return [x.strip() for x in s.split(";") if x.strip()]
    if "|" in s:
        return [x.strip() for x in s.split("|") if x.strip()]
    if "," in s and not s.startswith("E_"):
        return [x.strip() for x in s.split(",") if x.strip()]

    return [s]


def parse_maybe_obj(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    s = clean_text(value)
    if not s:
        return None

    try:
        return json.loads(s)
    except Exception:
        pass

    try:
        return ast.literal_eval(s)
    except Exception:
        return s


def load_json_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        print(f"[WARN] 找不到文件：{path}", file=sys.stderr)
        return {}

    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        return data

    print(f"[WARN] JSON 不是 dict 格式：{path}", file=sys.stderr)
    return {}


def extract_token_base_and_value(token: Any) -> Tuple[str, str]:
    """
    DDXPlus evidence token 可能长这样：
        E_53
        E_53_@_V_0
        E_53:V_0
    这里尽量拆出 evidence id 和取值。
    """
    s = clean_text(token).strip("'\"")
    if not s:
        return "", ""

    if "_@_" in s:
        base, val = s.split("_@_", 1)
        return base, val

    if "@" in s:
        base, val = s.split("@", 1)
        return base.strip("_:- "), val.strip("_:- ")

    if ":" in s and s.startswith("E_"):
        base, val = s.split(":", 1)
        return base, val

    return s, ""


def pick_first_string(d: Dict[str, Any], keys: List[str]) -> str:
    for k in keys:
        if k in d and d[k] not in (None, ""):
            extracted = text_from_any(d[k])
            if extracted:
                return extracted

    # 兼容 key 大小写/符号差异
    nd = {norm_key(k): v for k, v in d.items()}
    for k in keys:
        nk = norm_key(k)
        if nk in nd and nd[nk] not in (None, ""):
            extracted = text_from_any(nd[nk])
            if extracted:
                return extracted

    return ""


def evidence_token_to_text(token: Any, evidences: Dict[str, Any]) -> str:
    base, val = extract_token_base_and_value(token)
    if not base:
        return ""

    ev = evidences.get(base)

    # 有些 JSON 可能 key 不是标准大小写
    if ev is None:
        lower_map = {str(k).lower(): v for k, v in evidences.items()}
        ev = lower_map.get(base.lower())

    # 如果 val 本身是多语言 dict 字符串，先提取英文
    val_extracted = text_from_any(val)

    if isinstance(ev, dict):
        text = pick_first_string(
            ev,
            [
                "name",
                "label",
                "display_name",
                "question_en",
                "question",
                "question_text",
                "description",
                "meaning",
                "text",
                "code_question",
                "code",
                "id",
                "key",
            ],
        )

        # 如果常规字段没取到，扫描所有 key，尽量找英文问题/标签
        # 避免把 E_91 / code_question 这种编码本身当成症状文本。
        if not text or re.fullmatch(r"E_\d+", str(text).strip()):
            better_text = ""
            for k, v in ev.items():
                lk = str(k).lower()
                if any(word in lk for word in ["question", "name", "label", "desc", "meaning", "text", "en"]):
                    candidate = text_from_any(v)
                    if candidate and not re.fullmatch(r"E_\d+", candidate.strip()):
                        better_text = candidate
                        break
            if better_text:
                text = better_text

        # 取值含义，例如 V_0 / V_1 / V_104
        val_text = ""
        for values_key in [
            "possible-values",
            "possible_values",
            "values",
            "value_meaning",
            "value_meanings",
            "possible_answers",
            "possible_answers_en",
            "answers",
        ]:
            values = ev.get(values_key)
            if isinstance(values, dict):
                candidates = [val, val_extracted, str(val).upper(), str(val).lower()]
                if str(val).startswith("V_") or str(val).startswith("v_"):
                    candidates.append(str(val)[2:])
                for c in candidates:
                    if c in values:
                        val_text = text_from_any(values[c])
                        break
                    # 兼容大小写
                    for kk, vv in values.items():
                        if str(kk).lower() == str(c).lower():
                            val_text = text_from_any(vv)
                            break
                    if val_text:
                        break
            elif isinstance(values, list):
                # 如果 val 是 V_0，尝试拿 index 0
                m = re.search(r"(\d+)$", str(val))
                if m:
                    idx = int(m.group(1))
                    if 0 <= idx < len(values):
                        val_text = text_from_any(values[idx])
            if val_text:
                break

        # 有些证据值直接放在 ev 的 top-level 里，比如 values_en / data
        if not val_text and val:
            for k, v in ev.items():
                if str(k).lower() in [str(val).lower(), str(val_extracted).lower()]:
                    val_text = text_from_any(v)
                    break

        if not text:
            text = base

        if val and val_text:
            return f"{text}: {val_text}"
        if val and val_extracted and val_extracted != val:
            return f"{text}: {val_extracted}"
        if val:
            return f"{text}: {val}"
        return text

    if isinstance(ev, str):
        ev_text = text_from_any(ev)
        if val_extracted:
            return f"{ev_text}: {val_extracted}"
        return ev_text

    # 找不到映射时，至少保留原 token，避免症状空掉
    return text_from_any(token) or clean_text(token)


def build_symptom_text(row: Dict[str, Any], evidences: Dict[str, Any]) -> Tuple[str, str]:
    initial = get_field(row, ["INITIAL_EVIDENCE", "initial_evidence", "initialEvidence"])
    evidence_raw = get_field(row, ["EVIDENCES", "evidences", "evidence"])

    tokens: List[Any] = []
    if initial:
        tokens.append(initial)
    tokens.extend(parse_maybe_list(evidence_raw))

    seen = set()
    parts: List[str] = []
    for token in tokens:
        text = evidence_token_to_text(token, evidences)
        text = clean_text(text)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        parts.append(text)

    if not parts:
        # 极端情况下字段名不一致，做一个粗糙兜底
        for possible in ["SYMPTOMS", "symptoms", "DESCRIPTION", "description"]:
            v = get_field(row, [possible])
            if v:
                parts.append(v)

    symptom_text = "; ".join(parts)
    self_report = symptom_text
    return symptom_text, self_report


def format_differential(value: Any, max_items: int = 8) -> str:
    obj = parse_maybe_obj(value)

    if not obj:
        return ""

    items: List[str] = []

    if isinstance(obj, list):
        for x in obj[:max_items]:
            if isinstance(x, (list, tuple)) and len(x) >= 1:
                name = clean_text(x[0])
                prob = clean_text(x[1]) if len(x) > 1 else ""
                if name and prob:
                    items.append(f"{name} ({prob})")
                elif name:
                    items.append(name)
            elif isinstance(x, dict):
                name = pick_first_string(x, ["name", "condition", "pathology", "diagnosis"])
                prob = pick_first_string(x, ["probability", "score", "proba"])
                if name and prob:
                    items.append(f"{name} ({prob})")
                elif name:
                    items.append(name)
                else:
                    items.append(clean_text(x))
            else:
                items.append(clean_text(x))
    elif isinstance(obj, dict):
        # 可能是 {condition: probability}
        for i, (k, v) in enumerate(obj.items()):
            if i >= max_items:
                break
            if isinstance(v, (int, float, str)):
                items.append(f"{clean_text(k)} ({clean_text(v)})")
            else:
                items.append(clean_text(k))
    else:
        return clean_text(obj)

    return "; ".join([x for x in items if x])




def index_evidences(raw: Dict[str, Any]) -> Dict[str, Any]:
    """
    给 release_evidences.json 建一个更强的索引。

    有些 DDXPlus 版本可能不是：
        {"E_91": {...}}
    而是：
        {"some_name": {"code_question": "E_91", ...}}

    所以这里会递归扫描所有 dict/list，只要里面出现 E_数字，就把它映射到对应 dict。
    """
    indexed: Dict[str, Any] = {}

    def add_alias(alias: Any, obj: Any) -> None:
        s = clean_text(alias)
        if not s:
            return
        # 只提取 E_数字 这种证据 id
        for m in re.finditer(r"\bE_\d+\b", s):
            key = m.group(0)
            if key not in indexed:
                indexed[key] = obj
            indexed[key.lower()] = obj

    def walk(obj: Any) -> None:
        if isinstance(obj, dict):
            # 如果这个 dict 自己包含 E_xxx，就把整个 dict 作为该 E 的解释对象
            for k, v in obj.items():
                add_alias(k, obj)
                if isinstance(v, (str, int, float)):
                    add_alias(v, obj)

            for v in obj.values():
                walk(v)

        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    # 先保留原本的顶层 key 映射
    for k, v in raw.items():
        indexed[k] = v
        indexed[str(k).lower()] = v
        add_alias(k, v)

    walk(raw)

    print(f"[INFO] release_evidences 索引数量: {len(indexed)}")
    return indexed


def load_conditions_map(path: Path) -> Dict[str, str]:
    data = load_json_file(path)
    mapping: Dict[str, str] = {}

    for key, val in data.items():
        key_s = clean_text(key)
        display = key_s

        if isinstance(val, dict):
            display = pick_first_string(
                val,
                [
                    "condition_name",
                    "condition-name",
                    "condition_name_en",
                    "name",
                    "label",
                    "display_name",
                ],
            ) or key_s
        elif isinstance(val, str):
            display = clean_text(val)

        if key_s:
            mapping[key_s] = display
            mapping[key_s.lower()] = display
        if display:
            mapping[display] = display
            mapping[display.lower()] = display

    return mapping


def normalize_pathology(raw: str, conditions_map: Dict[str, str]) -> str:
    s = clean_text(raw)
    if not s:
        return ""

    return (
        conditions_map.get(s)
        or conditions_map.get(s.lower())
        or s.replace("_", " ").strip()
    )


def load_kg_mapping(xlsx_path: Path) -> Dict[str, Tuple[str, str]]:
    """
    从 dataset/knowledge graph of DDXPlus.xlsx 里尽量提取：
        disease/pathology/diagnosis -> (Level 2, Level 1)

    这个函数是“尽量兼容”，如果 Excel 表头和预期不同，会自动回退，不会影响主流程。
    """
    if not xlsx_path.exists():
        print(f"[WARN] 找不到 KG xlsx：{xlsx_path}，Level 1/2 会用兜底值。", file=sys.stderr)
        return {}

    try:
        from openpyxl import load_workbook
    except Exception:
        print("[WARN] 没有安装 openpyxl，跳过 xlsx 映射。可以执行：pip install openpyxl", file=sys.stderr)
        return {}

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        print(f"[WARN] 读取 KG xlsx 失败：{e}", file=sys.stderr)
        return {}

    rows = ws.iter_rows(values_only=True)
    try:
        header = [clean_text(x) for x in next(rows)]
    except StopIteration:
        return {}

    header_norm = [norm_key(h) for h in header]

    def find_col(candidates: List[str]) -> Optional[int]:
        cand_norm = [norm_key(c) for c in candidates]
        for i, h in enumerate(header_norm):
            if h in cand_norm:
                return i
        for i, h in enumerate(header_norm):
            for c in cand_norm:
                if c and c in h:
                    return i
        return None

    disease_col = find_col(["Disease", "Pathology", "Diagnosis", "Processed Diagnosis", "Condition", "Name"])
    level2_col = find_col(["Level 2", "Level2", "Subcategory", "Sub Category", "Sub-category", "L2"])
    level1_col = find_col(["Level 1", "Level1", "Category", "L1"])

    if disease_col is None:
        print("[WARN] KG xlsx 中没有识别到 Disease/Pathology/Diagnosis 列，跳过映射。", file=sys.stderr)
        return {}

    mapping: Dict[str, Tuple[str, str]] = {}

    for r in rows:
        if not r:
            continue
        disease = clean_text(r[disease_col] if disease_col < len(r) else "")
        if not disease:
            continue

        level2 = clean_text(r[level2_col] if level2_col is not None and level2_col < len(r) else "")
        level1 = clean_text(r[level1_col] if level1_col is not None and level1_col < len(r) else "")

        if not level2:
            level2 = disease
        if not level1:
            level1 = "DDXPlus"

        mapping[disease] = (level2, level1)
        mapping[disease.lower()] = (level2, level1)
        mapping[disease.replace("_", " ").lower()] = (level2, level1)

    print(f"[INFO] KG 映射读取完成：{len(mapping)} 个 key")
    return mapping


def find_split_file(raw_dir: Path, split: str) -> Optional[Path]:
    """
    优先找：
        release_train_patients.zip
        release_validate_patients.zip
        release_test_patients.zip
    也兼容 train.csv / validate.csv / test.csv
    """
    candidates = [
        raw_dir / f"release_{split}_patients.zip",
        raw_dir / f"release_{split}_patients.csv",
        raw_dir / f"{split}.csv",
        raw_dir / f"{split}_patients.csv",
    ]

    for p in candidates:
        if p.exists():
            return p

    # 宽松匹配
    patterns = [
        f"*{split}*patients*.zip",
        f"*{split}*patients*.csv",
        f"*{split}*.csv",
    ]
    for pat in patterns:
        matches = sorted(raw_dir.glob(pat))
        if matches:
            return matches[0]

    return None


def iter_csv_rows_from_file(path: Path) -> Iterator[Dict[str, Any]]:
    """
    兼容 DDXPlus 文件：
    - zip 里面是 csv
    - zip 里面是 json
    - zip 里面是 jsonl
    - zip 里面是没有后缀名的 csv/json/jsonl，例如 release_train_patients
    - 外部直接是 csv/json/jsonl
    """

    def rows_from_csv_binary(bf) -> Iterator[Dict[str, Any]]:
        text = io.TextIOWrapper(bf, encoding="utf-8-sig", newline="")
        reader = csv.DictReader(text)
        for row in reader:
            yield row

    def rows_from_json_binary(bf) -> Iterator[Dict[str, Any]]:
        data = json.load(io.TextIOWrapper(bf, encoding="utf-8-sig"))
        if isinstance(data, list):
            for item in data:
                if isinstance(item, dict):
                    yield item
        elif isinstance(data, dict):
            for key in ["patients", "data", "records", "items"]:
                if key in data and isinstance(data[key], list):
                    for item in data[key]:
                        if isinstance(item, dict):
                            yield item
                    return
            for _, item in data.items():
                if isinstance(item, dict):
                    yield item

    def rows_from_jsonl_binary(bf) -> Iterator[Dict[str, Any]]:
        text = io.TextIOWrapper(bf, encoding="utf-8-sig")
        for line in text:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            if isinstance(obj, dict):
                yield obj

    def guess_kind_from_bytes(data: bytes, name: str) -> str:
        """
        根据内容猜文件类型。
        DDXPlus 的 zip 里常见文件名没有 .csv 后缀，但内容其实是 CSV。
        """
        head = data[:4096].lstrip()
        if not head:
            raise FileNotFoundError(f"{name} 是空文件")

        # JSON / JSONL
        if head.startswith(b"{") or head.startswith(b"["):
            # 可能是普通 json，也可能是 jsonl；优先当普通 json 读，失败时外层会再试 jsonl
            return "json"

        # CSV：通常第一行包含 AGE,SEX,PATHOLOGY,EVIDENCES 等字段
        lower_head = head[:1000].decode("utf-8", errors="ignore").lower()
        csv_markers = ["age", "sex", "pathology", "evidences", "initial_evidence", "differential"]
        if "," in lower_head and any(m in lower_head for m in csv_markers):
            return "csv"

        # 兜底：DDXPlus extensionless patient 文件基本就是 CSV
        return "csv"

    def rows_from_bytes(data: bytes, name: str) -> Iterator[Dict[str, Any]]:
        lower = name.lower()

        if lower.endswith(".csv"):
            yield from rows_from_csv_binary(io.BytesIO(data))
            return

        if lower.endswith((".jsonl", ".ndjson")):
            yield from rows_from_jsonl_binary(io.BytesIO(data))
            return

        if lower.endswith(".json"):
            yield from rows_from_json_binary(io.BytesIO(data))
            return

        if lower.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data), "r") as inner:
                inner_names = [n for n in inner.namelist() if not n.endswith("/")]
                print(f"[INFO] {name} 内部嵌套文件列表前 20 个：")
                for n in inner_names[:20]:
                    print(f"       {n}")
                if not inner_names:
                    raise FileNotFoundError(f"{name} 的嵌套 zip 是空的")
                # 优先 patient/data 文件
                inner_names.sort(key=lambda n: (0 if "patient" in n.lower() else 1, n))
                with inner.open(inner_names[0], "r") as bf:
                    yield from rows_from_bytes(bf.read(), inner_names[0])
            return

        # 无后缀文件：按内容猜
        kind = guess_kind_from_bytes(data, name)
        if kind == "csv":
            yield from rows_from_csv_binary(io.BytesIO(data))
            return

        if kind == "json":
            try:
                yield from rows_from_json_binary(io.BytesIO(data))
                return
            except Exception:
                yield from rows_from_jsonl_binary(io.BytesIO(data))
                return

        raise FileNotFoundError(f"无法识别文件类型：{name}")

    suffix = path.suffix.lower()

    if suffix == ".zip":
        with zipfile.ZipFile(path, "r") as zf:
            names = [n for n in zf.namelist() if not n.endswith("/")]

            print(f"[INFO] {path.name} 内部文件列表前 20 个：")
            for n in names[:20]:
                print(f"       {n}")

            if not names:
                raise FileNotFoundError(f"{path} 是空 zip")

            # 优先 csv/json/jsonl，其次无后缀 patient 文件，最后嵌套 zip
            def priority(n: str) -> Tuple[int, str]:
                ln = n.lower()
                if ln.endswith(".csv"):
                    return (0, n)
                if ln.endswith((".jsonl", ".ndjson")):
                    return (1, n)
                if ln.endswith(".json"):
                    return (2, n)
                if "patient" in ln:
                    return (3, n)
                if ln.endswith(".zip"):
                    return (4, n)
                return (5, n)

            names.sort(key=priority)
            with zf.open(names[0], "r") as bf:
                yield from rows_from_bytes(bf.read(), names[0])
            return

    if suffix in [".csv", ".json", ".jsonl", ".ndjson", ".zip"]:
        with path.open("rb") as bf:
            yield from rows_from_bytes(bf.read(), path.name)
        return

    # 外部也可能是无后缀文件
    with path.open("rb") as bf:
        yield from rows_from_bytes(bf.read(), path.name)



def make_medrag_record(
    row: Dict[str, Any],
    participant_no: int,
    evidences: Dict[str, Any],
    conditions_map: Dict[str, str],
    kg_mapping: Dict[str, Tuple[str, str]],
) -> Dict[str, str]:
    age = get_field(row, ["AGE", "Age", "age"])
    sex = get_field(row, ["SEX", "Sex", "sex", "GENDER", "gender"])

    raw_pathology = get_field(row, ["PATHOLOGY", "Pathology", "pathology", "DIAGNOSIS", "diagnosis"])
    diagnosis = normalize_pathology(raw_pathology, conditions_map)

    symptom_text, self_report = build_symptom_text(row, evidences)

    diff_raw = get_field(row, ["DIFFERENTIAL_DIAGNOSIS", "differential_diagnosis", "DIFFERENTIAL", "differential"])
    diff_text = format_differential(diff_raw)

    level2, level1 = kg_mapping.get(diagnosis) or kg_mapping.get(diagnosis.lower()) or ("", "")
    if not level2:
        level2 = diagnosis or "Unknown"
    if not level1:
        level1 = "DDXPlus"

    return {
        "Participant No.": str(participant_no),
        "Age": age,
        "Sex": sex,
        "Pain Presentation and Description": symptom_text,
        "Pain descriptions and assorted symptoms (self-report)": self_report,
        "Pain restriction": diff_text,
        "Processed Diagnosis": diagnosis,
        "Level 2": level2,
        "Level 1": level1,
    }


def make_json_record(csv_record: Dict[str, str]) -> Dict[str, Any]:
    age = csv_record.get("Age", "")
    sex = csv_record.get("Sex", "")
    symptoms = csv_record.get("Pain Presentation and Description", "")
    diagnosis = csv_record.get("Processed Diagnosis", "")
    diff = csv_record.get("Pain restriction", "")

    text_parts = [
        f"Age: {age}" if age else "",
        f"Sex: {sex}" if sex else "",
        f"Symptoms: {symptoms}" if symptoms else "",
        f"Differential diagnosis: {diff}" if diff else "",
        f"Diagnosis: {diagnosis}" if diagnosis else "",
    ]
    text = "\n".join([p for p in text_parts if p])

    return {
        "Participant No.": int(csv_record.get("Participant No.", "0") or 0),
        "Age": age,
        "Sex": sex,
        "Symptoms": symptoms,
        "Diagnosis": diagnosis,
        "Differential Diagnosis": diff,
        "Level 2": csv_record.get("Level 2", ""),
        "Level 1": csv_record.get("Level 1", ""),
        "Text": text,
    }


def clear_participant_jsons(out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    for p in out_dir.glob("participant_*.json"):
        p.unlink()


def write_json_record(out_dir: Path, index: int, record: Dict[str, str]) -> None:
    obj = make_json_record(record)
    out_path = out_dir / f"participant_{index}.json"
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def process_split(
    split: str,
    raw_dir: Path,
    csv_writer: csv.DictWriter,
    global_start_no: int,
    json_out_dir: Optional[Path],
    max_rows: Optional[int],
    evidences: Dict[str, Any],
    conditions_map: Dict[str, str],
    kg_mapping: Dict[str, Tuple[str, str]],
) -> Tuple[int, int]:
    """
    返回：
        csv 写入数量, json 写入数量
    """
    split_file = find_split_file(raw_dir, split)
    if not split_file:
        print(f"[WARN] 没找到 {split} 数据文件，跳过。目录：{raw_dir}", file=sys.stderr)
        return 0, 0

    print(f"[INFO] 读取 {split}: {split_file}")

    csv_count = 0
    json_count = 0

    if json_out_dir is not None:
        json_out_dir.mkdir(parents=True, exist_ok=True)

    for row in iter_csv_rows_from_file(split_file):
        if max_rows is not None and csv_count >= max_rows:
            break

        participant_no = global_start_no + csv_count
        medrag_record = make_medrag_record(
            row=row,
            participant_no=participant_no,
            evidences=evidences,
            conditions_map=conditions_map,
            kg_mapping=kg_mapping,
        )

        # 如果症状和诊断都空，基本说明字段没读对，跳过
        if not medrag_record["Pain Presentation and Description"] and not medrag_record["Processed Diagnosis"]:
            continue

        csv_writer.writerow(medrag_record)
        csv_count += 1

        if json_out_dir is not None:
            json_count += 1
            write_json_record(json_out_dir, json_count, medrag_record)

    return csv_count, json_count


def main() -> None:
    root = repo_root()

    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-dir", default=str(root / "dataset" / "ddxplus_raw"))
    parser.add_argument("--max-train", type=int, default=1000, help="默认只取 1000 条训练样本用于先跑通")
    parser.add_argument("--max-validate", type=int, default=300, help="默认只取 300 条验证样本写入总 CSV")
    parser.add_argument("--max-test", type=int, default=300, help="默认只取 300 条测试样本用于先跑通")
    parser.add_argument("--all", action="store_true", help="全量生成，不限制条数。可能很慢，占用很多磁盘。")
    parser.add_argument("--no-clean", action="store_true", help="不清理旧 participant_*.json 文件")
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir).resolve()
    dataset_dir = root / "dataset"
    df_train_dir = dataset_dir / "df" / "train"
    df_test_dir = dataset_dir / "df" / "test"
    output_csv = dataset_dir / "AI Data Set with Categories.csv"

    if not raw_dir.exists():
        raise FileNotFoundError(
            f"找不到 DDXPlus 原始数据目录：{raw_dir}\n"
            f"请先把 release_train_patients.zip 等文件放到 dataset/ddxplus_raw/"
        )

    max_train = None if args.all else args.max_train
    max_validate = None if args.all else args.max_validate
    max_test = None if args.all else args.max_test

    dataset_dir.mkdir(parents=True, exist_ok=True)
    df_train_dir.mkdir(parents=True, exist_ok=True)
    df_test_dir.mkdir(parents=True, exist_ok=True)

    if not args.no_clean:
        print("[INFO] 清理旧的 participant_*.json 文件")
        clear_participant_jsons(df_train_dir)
        clear_participant_jsons(df_test_dir)

    evidences = index_evidences(load_json_file(raw_dir / "release_evidences.json"))
    conditions_map = load_conditions_map(raw_dir / "release_conditions.json")
    kg_mapping = load_kg_mapping(dataset_dir / "knowledge graph of DDXPlus.xlsx")

    total_csv = 0
    train_json = 0
    test_json = 0

    with output_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_CSV_COLUMNS)
        writer.writeheader()

        n_csv, n_json = process_split(
            split="train",
            raw_dir=raw_dir,
            csv_writer=writer,
            global_start_no=total_csv + 1,
            json_out_dir=df_train_dir,
            max_rows=max_train,
            evidences=evidences,
            conditions_map=conditions_map,
            kg_mapping=kg_mapping,
        )
        total_csv += n_csv
        train_json += n_json

        # validate 只写入总 CSV，不生成 json。MedRAG 当前你要的是 df/train 和 df/test。
        n_csv, _ = process_split(
            split="validate",
            raw_dir=raw_dir,
            csv_writer=writer,
            global_start_no=total_csv + 1,
            json_out_dir=None,
            max_rows=max_validate,
            evidences=evidences,
            conditions_map=conditions_map,
            kg_mapping=kg_mapping,
        )
        total_csv += n_csv

        n_csv, n_json = process_split(
            split="test",
            raw_dir=raw_dir,
            csv_writer=writer,
            global_start_no=total_csv + 1,
            json_out_dir=df_test_dir,
            max_rows=max_test,
            evidences=evidences,
            conditions_map=conditions_map,
            kg_mapping=kg_mapping,
        )
        total_csv += n_csv
        test_json += n_json

    print("\n[DONE] DDXPlus -> MedRAG 数据转换完成")
    print(f"[DONE] CSV: {output_csv}")
    print(f"[DONE] CSV 行数: {total_csv}")
    print(f"[DONE] train json 数量: {train_json} -> {df_train_dir}")
    print(f"[DONE] test json 数量: {test_json} -> {df_test_dir}")
    print("\n下一步可以先检查：")
    print(f"  dir \"{df_train_dir}\"")
    print(f"  dir \"{df_test_dir}\"")
    print(f"  python -c \"import pandas as pd; print(pd.read_csv(r'{output_csv}').head())\"")


if __name__ == "__main__":
    main()
